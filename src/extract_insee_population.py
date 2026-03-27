from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(__file__).with_name("Estimation_popu_2025_dpt_sexe_classe_age.xlsx")
OUTPUT_PATH = Path(__file__).with_name("insee_population_departements_wide_2022_2025.csv")
YEARS = ["2022", "2023", "2024", "2025"]

OUTPUT_COLUMNS = [
    "year",
    "department_code",
    "department_name",
    "ensemble_0_19",
    "ensemble_20_39",
    "ensemble_40_59",
    "ensemble_60_74",
    "ensemble_75_plus",
    "ensemble_total",
    "hommes_0_19",
    "hommes_20_39",
    "hommes_40_59",
    "hommes_60_74",
    "hommes_75_plus",
    "hommes_total",
    "femmes_0_19",
    "femmes_20_39",
    "femmes_40_59",
    "femmes_60_74",
    "femmes_75_plus",
    "femmes_total",
]


def iter_sheet_rows(sheet_name: str) -> list[dict[str, int | str]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    rows: list[dict[str, int | str]] = []

    for row in worksheet.iter_rows(min_row=6, values_only=True):
        department_code = row[0]
        department_name = row[1]

        if department_code is None or department_name is None:
            continue

        rows.append(
            {
                "year": int(sheet_name),
                "department_code": str(department_code).zfill(2),
                "department_name": str(department_name).strip(),
                "ensemble_0_19": int(row[2]),
                "ensemble_20_39": int(row[3]),
                "ensemble_40_59": int(row[4]),
                "ensemble_60_74": int(row[5]),
                "ensemble_75_plus": int(row[6]),
                "ensemble_total": int(row[7]),
                "hommes_0_19": int(row[8]),
                "hommes_20_39": int(row[9]),
                "hommes_40_59": int(row[10]),
                "hommes_60_74": int(row[11]),
                "hommes_75_plus": int(row[12]),
                "hommes_total": int(row[13]),
                "femmes_0_19": int(row[14]),
                "femmes_20_39": int(row[15]),
                "femmes_40_59": int(row[16]),
                "femmes_60_74": int(row[17]),
                "femmes_75_plus": int(row[18]),
                "femmes_total": int(row[19]),
            }
        )

    workbook.close()
    return rows


def main() -> None:
    extracted_rows: list[dict[str, int | str]] = []
    for year in YEARS:
        extracted_rows.extend(iter_sheet_rows(year))

    with OUTPUT_PATH.open("w", encoding="utf-8", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(extracted_rows)

    print(f"Wrote {len(extracted_rows)} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()